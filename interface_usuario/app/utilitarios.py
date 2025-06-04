import os
import openpyxl
import pyodbc
import requests
from requests.auth import HTTPDigestAuth
import shutil

def trigger_manual_correction(ip):

    # Configurações do dispositivo
    url = f"http://{ip}/ISAPI/Image/channels/2/ManualShutterCorrect"
    username = "admin"  # Substitua pelo seu nome de usuário
    password = "czcz8910"  # Substitua pela sua senha

    try:
        # Fazendo a requisição PUT com autenticação Digest
        response = requests.put(
            url,
            auth=HTTPDigestAuth(username, password),
            timeout=10  # Timeout de 10 segundos
        )

        # Verificando o status da resposta
        if response.status_code == 200:
            print("Correção manual disparada com sucesso!")
            print("Resposta:", response.text)  # Exibe o corpo da resposta (pode ser um JSON)
        else:
            print(f"Falha ao disparar a correção manual. Status Code: {response.status_code}")
            print("Resposta:", response.text)

    except requests.exceptions.RequestException as e:
        print(f"Erro ao fazer a requisição: {e}")

def copiar_para_rede(caminho_local, caminho_rede):
    try:
        if not os.path.exists(caminho_rede):
            os.makedirs(caminho_rede)  # Cria o diretório de rede se não existir
        
        arquivo_nome = os.path.basename(caminho_local)
        destino = os.path.join(caminho_rede, arquivo_nome)
        
        shutil.copy2(caminho_local, destino)
        print(f"Arquivo copiado para {destino}")
    
    except Exception as e:
        print(f"Erro ao copiar arquivo para a rede: {e}")


def load_config(config_file):

    config = {}
    try:
        with open(config_file, 'r') as f:
            for line in f:
                line = line.strip()
                if line and '=' in line:
                    key, value = line.split('=', 1)
                    config[key.strip()] = value.strip()
        return config

    except FileNotFoundError:
        print(f"Erro: arquivo de configuração não encontrado em {config_file}")
        raise
    except Exception as e:
        print(f"Erro ao carregar configuração: {e}")
        raise


def consultar_placas(data_abate):
    #Consulta o banco SQL Server e retorna os dados das placas com a estrutura correta
    try:
        conn = pyodbc.connect(
            "DRIVER={ODBC Driver 18 for SQL Server};"
            "SERVER=SRVBMAIS;"
            "DATABASE=BM;"
            "UID=BM;"
            "PWD=bm0941;"
            "TrustServerCertificate=Yes",
            timeout=5
        )
        cursor = conn.cursor()

        query = """
            SELECT
                OEPAI.COD_INC AS COD_INC_PAI, 
                OEPAI.PLACA,  
                CONCAT(OEPAI.cod_pedcom, '-', OEPAI.COD_ORDEM) AS ORDEM, 
                ISNULL(gta.num_gta, '') AS GTA, 
                ISNULL((Qtde_M_0_12 + Qtde_F_0_12 + Qtde_M_13_24 + Qtde_F_13_24 + 
                        Qtde_M_25_36 + Qtde_F_25_36 + Qtde_M_37_99 + Qtde_F_37_99), '') AS QUANT,
                ISNULL(gta.nome_forn_imp, '') AS FORNECEDOR, 
                ISNULL(gta.nome_faz_imp, '') AS FAZENDA, 
                CASE 
                    WHEN gta.municipio_faz_imp IS NULL THEN '' 
                    ELSE CONCAT(gta.municipio_faz_imp, ' / ', gta.uf_faz_imp) 
                END AS MUNICIPIO,
                CASE 
                    WHEN r.cod_rampa = '001' THEN 'P1'
                    WHEN r.cod_rampa = '002' THEN 'P5'
                    ELSE ISNULL(r.cod_rampa, '')
                END AS COD_RAMPA
            FROM APIS.dbo.t_ordem_gado oepai WITH (NOLOCK)
            INNER JOIN APIS.dbo.t_pedcompbov pedc WITH (NOLOCK) 
                ON oepai.cod_pedcom = pedc.cod_pedcom 
            LEFT JOIN APIS.dbo.t_ordem_retorno_gta gta WITH (NOLOCK) 
                ON oepai.cod_inc = gta.cod_inc_pai
            LEFT JOIN APIS.dbo.t_ordem_retorno r WITH (NOLOCK)
                ON oepai.cod_inc = r.cod_inc_pai
            WHERE pedc.abate = ?
            ORDER BY OEPAI.PLACA
        """

        cursor.execute(query, (data_abate,))
        resultados = cursor.fetchall()
        conn.close()

        if not resultados:
            print(f"Nenhuma placa encontrada para {data_abate}. Solicitando inserção manual.")
            return [("N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A")]
        return resultados

    except pyodbc.Error as e:
        print(f"Erro de conexão com o SQL Server: {e}")
        return [("N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A", "N.A")]

def consultar_resultados_excel(caminho_excel, placa, ordem):
    """Consulta a planilha RESULTADOS.xlsx e retorna os resultados para a placa e ordem especificadas."""
    try:
        wb = openpyxl.load_workbook(caminho_excel)
        ws = wb.active
        resultados = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) >= 7:  # Data, Hora, Placa, Ordem, Sequencial, Quantidade, Caminho
                row_data, row_hora, row_placa, row_ordem, row_sequencial, row_quantidade, row_caminho = row[:7]
                if row_placa == placa.replace("-", "").strip() and row_ordem == ordem.replace("-", "").strip():
                    resultados.append((row_hora, row_placa, row_ordem, row_sequencial, row_quantidade, row_caminho))
            else:
                print(f"LINHA {row}NÃO ENTROU NO IF")
        wb.close()
        print(f"RESULTADOS = {resultados}")
        return resultados
    except FileNotFoundError:
        print(f"Arquivo {caminho_excel} não encontrado. Nenhum resultado retornado.")
        return []
    except Exception as e:
        print(f"Erro ao consultar planilha: {e}")
        return []
    
def registrar_resultado(data, placa, sequencial, quantidade, caminho_excel, hora, ordem_compra):
    """Registra os resultados no Excel (futuramente substituível por INSERT no banco)."""
    output_excel = caminho_excel
    if not os.path.exists(caminho_excel):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Data", "Hora", "Placa", "Ordem", "Sequencial", "Quantidade", "Caminho"])
    else:
        wb = openpyxl.load_workbook(caminho_excel)
        ws = wb.active
    ws.append([data, hora, placa, ordem_compra, int(sequencial), quantidade])
    wb.save(output_excel)
    print(f"Resultado salvo em: {output_excel}")

